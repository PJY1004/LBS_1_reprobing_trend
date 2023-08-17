# 따로 이미지 나오지 않고 완성된 excel 파일이 나오도록 함!!

import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles import Border, Side
from openpyxl.styles import PatternFill
from openpyxl.drawing.image import Image
from datetime import datetime, timedelta
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
import os

reprobing_trend_file_name = (datetime.today()-timedelta(days=1)).strftime("%Y%m%d")[2:]+'_REPROBING_TREND.xlsx'

df = pd.read_excel('remon.xlsx') # 같은 폴더에 있는 'remon.xlsx' 파일을 읽음!!
df = df.sort_values('PARTID',ascending = False)    # 'PARTID' 기준 오름차순 정렬

partid_list = sorted(list(set(df['PARTID'])),reverse=True)

wb = openpyxl.Workbook(reprobing_trend_file_name)
wb.save(reprobing_trend_file_name)  # wb.openpyxl.Workbook ~ wb.save 가 한 문장 단위라서 다시 엑셀 편집할 땐 wb.openpyxl.Workbook로 시작해야함..

for partid in partid_list:
    LD = df[df['PARTID'] == partid]       # 읽은 데이터 중 'PARTID'가 a인 record만 걸림..

    # 최종 차트 + 환원율 Table 구현이 main 목표!!
    ############################
    # (1) PASS 데이터 추출 + Fail 데이터 추출 (before1 , after1, qty 에 대해서만..)
    LD_PASS = LD[LD['AFTER1'] == 1]                   # 'AFTER1'이 1인 record만 LD_PASS에 저장
    LD_PASS = LD_PASS[['BEFORE1', 'AFTER1', 'QTY']]   # LD_PASS 데이터 중 'BEFORE1', 'AFTER1', 'QTY' 열만 저장
    LD_FAIL = LD                                      # LD_FAIL엔 전체 record 저장
    LD_FAIL = LD_FAIL[['BEFORE1', 'AFTER1', 'QTY']]   # LD_FAIL 데이터 중 'BEFORE1', 'AFTER1', 'QTY' 열만 저장

    LD_PASS = LD_PASS.sort_values('BEFORE1')  # BEFORE1 에 대해 오름차순 정렬
    LD_FAIL = LD_FAIL.sort_values('BEFORE1')

    BEFORE_LD_PASS = sorted(set(LD_PASS['BEFORE1']))  # 딕셔너리 -> 리스트 자료형 변환.
    BEFORE_LD_FAIL = sorted(set(LD_FAIL['BEFORE1']))  # "BEFORE1" 중복값 제거한 오름차순 리스트 저장

    #########################

    # (2) pass/fail 에 대해 before1로 구분해서 sum(qty) 각각 구하기
    # BEFORE1마다 SUM 구해 pivot table 형식으로 데이터프레임 만들기
    SUM_LD_PASS = []
    SUM_LD_FAIL = []

    for i in BEFORE_LD_FAIL:  # BEFORE1 항목 단위로 SUM값 구하기
        SUM_LD_PASS.append(LD_PASS['QTY'][LD_PASS['BEFORE1'] == i].sum()) # before1 이 i인 LD_PASS 레코드 중
                                                                # QTY열에 해당하는 데이터의 총합을 리스트 SUM_LD_PASS에 넣음
    for i in BEFORE_LD_FAIL:
        SUM_LD_FAIL.append(LD_FAIL['QTY'][LD_FAIL['BEFORE1'] == i].sum())

    #################

    # (3) Bin No , sum(qty) 로 피벗 테이블 제작
    PIVOT_LD = pd.DataFrame({'FAIL 합계': SUM_LD_FAIL, 'PASS 합계': SUM_LD_PASS}, index=BEFORE_LD_FAIL)

    ## 환원율 계산
    PIVOT_LD['환원율'] = round(PIVOT_LD['PASS 합계'] / PIVOT_LD['FAIL 합계']*100, 2)

    #################

    # (4) 차트 그리기
    fig1 = plt.subplots(1)
    bar_width = 0.3
    index = np.arange(len(PIVOT_LD.index))

    b1 = plt.bar(index + 0.15, PIVOT_LD['FAIL 합계'], bar_width, alpha=0.5, color='blue', label='FAIL')
    b2 = plt.bar(index + 0.45, PIVOT_LD['PASS 합계'], bar_width, alpha=0.5, color='red', label='PASS')
    plt.xticks(np.arange(bar_width, len(PIVOT_LD.index) + bar_width, 1), PIVOT_LD.index)
    plt.title(partid + " REPROBING PASS RECOVERY TREND")
    plt.legend(loc='upper right')
    plt.savefig(partid[3:6] + '.png')

    #################

    # (5)엑셀에 쓰기
    PIVOT_LD['환원율'] = PIVOT_LD['환원율'].astype(str) + '%'


    # with pd.ExcelWriter(partid[3:6] + ' Reprobing Trend.xlsx') as writer:  # EX) a = 'S5P9855F01-N2Z' => 98
    with pd.ExcelWriter(reprobing_trend_file_name, mode='a', engine='openpyxl') as writer:
        PIVOT_LD.to_excel(writer, sheet_name=partid[3:6] + ' 피벗')   #     a = 'S5KHP2SX03-Y1Z' => HP2 로 문자따와서 엑셀로 저장
        LD_PASS.to_excel(writer, sheet_name=partid[3:6] + ' PASS')
        LD_FAIL.to_excel(writer, sheet_name=partid[3:6] + ' FAIL')

    ## 엑셀 FONT, CELL 서식 변경 ##
    wb = load_workbook(reprobing_trend_file_name)  # 엑셀 파일 다시 가져와서
    try:
        wb.remove(wb['Sheet'])
    except:
        pass
    ws = wb[partid[3:6] + ' 피벗']  # 활성화

    ws['A1'].fill = PatternFill(start_color='9BC2E6', fill_type='solid')
    ws['B1'].fill = PatternFill(start_color='9BC2E6', fill_type='solid')
    ws['C1'].fill = PatternFill(start_color='9BC2E6', fill_type='solid')
    ws['D1'].fill = PatternFill(start_color='9BC2E6', fill_type='solid')

    ws['A1'].value = "BIN"
    ws['B1'].value = "FAIL"
    ws['C1'].value = "PASS"

    for data in ws['A']:
        data.alignment = Alignment(horizontal="center")

    for data in ws['B']:
        data.font = Font(bold=True)
        data.border = Border(Side('thin'), Side('thin'), Side('thin'), Side('thin'))
        data.alignment = Alignment(horizontal="center")

    for data in ws['C']:
        data.font = Font(bold=True)
        data.border = Border(Side('thin'), Side('thin'), Side('thin'), Side('thin'))
        data.alignment = Alignment(horizontal="center")

    for data in ws['D']:
        data.font = Font(bold=True)
        data.border = Border(Side('thin'), Side('thin'), Side('thin'), Side('thin'))
        data.alignment = Alignment(horizontal="center")

    ws['A1'].font = Font(bold=True)
    ws["A1"].border = Border(Side('thin'), Side('thin'), Side('thin'), Side('thin'))

    total_record = ws.max_row
    image = Image(partid[3:6] + '.png')
    ws.add_image(image,'E'+str(total_record+1))  ## 이미지 삽입

    wb.save(filename=reprobing_trend_file_name)
    os.remove(partid[3:6] + '.png')

wb = load_workbook(reprobing_trend_file_name)  # 엑셀 파일 다시 가져와서
count_worksheet = 0
for partid in partid_list:
    ws = wb[partid[3:6] + ' 피벗']  # 활성화
    wb.move_sheet(ws,-2*count_worksheet)
    count_worksheet += 1
wb.save(reprobing_trend_file_name)

# for partid in partid_list:
#     wb = load_workbook(partid[3:6] + ' Reprobing Trend.xlsx')  # 엑셀 파일 다시 가져와서
#     ws = wb.active  # 활성화
#################
